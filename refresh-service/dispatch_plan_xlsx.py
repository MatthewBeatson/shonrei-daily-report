"""Monthly Dispatch Plan -- workbook writer.

Renders the {month_key: {week: [Group]}} schedule from dispatch_plan_schedule
into an .xlsx matching the real "<MONTH> ORDERS PLAN" reference workbook's
layout cell-for-cell (confirmed by inspecting the real June 2026 file this
project replaces -- see the plan doc, not guessed), plus a new first
"Overview" tab (Matthew's call, not in the original workbook) rolling up
every covered month.

Every total is a live Excel formula, never a hardcoded Python-computed
number -- same convention as the reference workbook itself.
"""
from __future__ import annotations
import calendar
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from dispatch_plan_schedule import week_ranges_for_month, month_from_key

CURRENCY_FMT = '_-"$"* #,##0_-;\\-"$"* #,##0_-;_-"$"* "-"??_-;_-@_-'
DATE_FMT = '[$-F800]dddd\\,\\ mmmm\\ dd\\,\\ yyyy'
BOLD = Font(bold=True)
WEEK_HEADER_FILL = PatternFill('solid', fgColor='D9D9D9')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
GREEN_FONT = Font(bold=True, color='006100')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
RED_FONT = Font(bold=True, color='9C0006')


def _month_label(month_key: str) -> str:
    y, m = month_from_key(month_key)
    return calendar.month_name[m].upper()


def _sheet_title(month_key: str, used_titles: set[str]) -> str:
    y, m = month_from_key(month_key)
    title = f'{_month_label(month_key)} ORDERS PLAN'
    if title in used_titles:
        title = f'{_month_label(month_key)} {y} ORDERS PLAN'
    used_titles.add(title)
    return title[:31]  # Excel sheet name limit


def _date_range_label(start: date, end: date) -> str:
    month_name = calendar.month_name[start.month]
    if start.day == end.day:
        return f'{start.day} {month_name}'
    return f'{start.day}-{end.day} {month_name}'


def _write_month_sheet(
    ws: Worksheet,
    month_key: str,
    weeks: dict[int, list],
    monthly_target: float,
    invoiced_to_date: float,
    breakeven: float | None,
    as_at: date,
) -> dict:
    """Returns the address (within this sheet) of the cells the Overview
    sheet needs to reference: target, invoiced, total_projected, breakeven."""
    y, m = month_from_key(month_key)
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 14

    ws['B1'] = f'{_month_label(month_key)} ORDERS PLAN'
    ws['B1'].font = BOLD
    ws['E1'] = "THIS MONTH'S TARGET:"
    ws['E1'].font = BOLD
    ws['F1'] = monthly_target
    ws['F1'].font = BOLD
    ws['F1'].number_format = CURRENCY_FMT

    ws['D2'] = 'SO#'
    ws['E2'] = 'DISPATCH DATE'
    ws['F2'] = 'ORDER VALUE'
    for cell in ('D2', 'E2', 'F2'):
        ws[cell].font = BOLD

    row = 3
    prev_cumulative_cell = None  # week 1's base is the Invoiced sales cell, set below via a forward reference
    week_end_cells = []

    all_week_numbers = sorted(week_ranges_for_month(y, m), key=lambda w: w[0])
    week_lookup = {w[0]: (w[1], w[2]) for w in all_week_numbers}

    for week_no in sorted(weeks.keys()):
        start, end = week_lookup.get(week_no, (None, None))
        ws.cell(row=row, column=2, value=f'WEEK {week_no} DISPATCH').font = BOLD
        for col in range(2, 7):
            ws.cell(row=row, column=col).fill = WEEK_HEADER_FILL
        row += 1
        if start and end:
            ws.cell(row=row, column=2, value=_date_range_label(start, end))
        row += 1

        first_order_row = row
        for g in sorted(weeks[week_no], key=lambda g: (g.scheduling_date, g.customer)):
            ws.cell(row=row, column=2, value=g.customer)
            ws.cell(row=row, column=3, value=g.label)
            ws.cell(row=row, column=4, value=', '.join(g.order_numbers))
            if g.scheduling_date:
                c = ws.cell(row=row, column=5, value=datetime(g.scheduling_date.year, g.scheduling_date.month, g.scheduling_date.day))
                c.font = BOLD
                c.number_format = DATE_FMT
            c = ws.cell(row=row, column=6, value=g.value_excl_gst)
            c.number_format = CURRENCY_FMT
            row += 1
        last_order_row = row - 1

        week_sum_row = row
        ws.cell(row=row, column=2, value=f'BUDGETED SALES FOR WEEK {week_no}:').font = BOLD
        sum_formula = f'=SUM(F{first_order_row}:F{last_order_row})' if last_order_row >= first_order_row else 0
        c = ws.cell(row=row, column=3, value=sum_formula)
        c.font = BOLD
        c.number_format = CURRENCY_FMT

        cumulative_row = row
        ws.cell(row=row, column=5, value=f'BUDGETED SALES AT END OF WEEK {week_no}:').font = BOLD
        base = prev_cumulative_cell if prev_cumulative_cell else 'INVOICED_SALES_CELL'
        cumulative_formula = f'=C{week_sum_row}+{base}'
        c = ws.cell(row=row, column=6, value=cumulative_formula)  # base placeholder patched below
        c.font = BOLD
        c.number_format = CURRENCY_FMT
        week_end_cells.append(f'F{cumulative_row}')
        prev_cumulative_cell = f'F{cumulative_row}'
        row += 2  # blank separator row before next week header

    invoiced_row = row
    ws.cell(row=row, column=3, value='AS AT:')
    ws.cell(row=row, column=4, value=datetime(as_at.year, as_at.month, as_at.day)).number_format = 'mm-dd-yy'
    ws.cell(row=row, column=5, value='Invoiced sales')
    c = ws.cell(row=row, column=6, value=invoiced_to_date)
    c.font = BOLD
    c.number_format = CURRENCY_FMT
    row += 1

    # Patch every week's "base" placeholder now that we know the invoiced-
    # sales cell's real address (week 1 needed a forward reference to a row
    # that didn't exist yet when it was written).
    invoiced_cell = f'F{invoiced_row}'
    for r in range(3, row):
        cell = ws.cell(row=r, column=6)
        if isinstance(cell.value, str) and 'INVOICED_SALES_CELL' in cell.value:
            cell.value = cell.value.replace('INVOICED_SALES_CELL', invoiced_cell)

    total_row = row
    ws.cell(row=row, column=5, value='TOTAL PROJECTED SALES').font = BOLD
    last_week_end = week_end_cells[-1] if week_end_cells else invoiced_cell
    c = ws.cell(row=row, column=6, value=f'={last_week_end}')
    c.font = GREEN_FONT
    c.fill = GREEN_FILL
    c.number_format = CURRENCY_FMT
    row += 1

    breakeven_row = row
    ws.cell(row=row, column=5, value='Breakeven (survival)').font = BOLD
    c = ws.cell(row=row, column=6, value=breakeven if breakeven is not None else 0)
    c.number_format = CURRENCY_FMT
    if breakeven is None:
        ws.cell(row=row, column=7, value='Not set in reporting.manual_inputs -- edit on the daily dashboard')
    row += 1

    ws.cell(row=row, column=5, value='Profit / loss').font = BOLD
    c = ws.cell(row=row, column=6, value=f'=F{total_row}-F{breakeven_row}')
    c.number_format = CURRENCY_FMT
    row += 2

    return {
        'sheet_title': ws.title,
        'target_cell': 'F1',
        'invoiced_cell': invoiced_cell,
        'total_projected_cell': f'F{total_row}',
        'breakeven_cell': f'F{breakeven_row}',
    }


def _write_holding_section(ws: Worksheet, start_row: int, holding: list) -> int:
    row = start_row
    ws.cell(row=row, column=2, value="Other orders in system, but can't be sent/invoiced until later:")
    for col in range(2, 5):
        cell = ws.cell(row=row, column=col)
        cell.font = RED_FONT
        cell.fill = RED_FILL
    row += 1

    first_row = row
    for g in sorted(holding, key=lambda g: (g.customer, g.order_date or date.min)):
        ws.cell(row=row, column=2, value=g.customer)
        ws.cell(row=row, column=3, value=g.label)
        ws.cell(row=row, column=4, value=', '.join(g.order_numbers))
        c = ws.cell(row=row, column=6, value=g.value_excl_gst)
        c.number_format = CURRENCY_FMT
        row += 1
    last_row = row - 1

    ws.cell(row=row, column=2, value='TOTAL').font = RED_FONT
    ws.cell(row=row, column=2).fill = RED_FILL
    formula = f'=SUM(F{first_row}:F{last_row})' if last_row >= first_row else 0
    c = ws.cell(row=row, column=6, value=formula)
    c.font = RED_FONT
    c.fill = RED_FILL
    c.number_format = CURRENCY_FMT
    return row + 1


def build_workbook(
    schedule: dict[str, dict[int, list]],
    holding: list,
    monthly_target: float,
    invoiced_to_date_by_month: dict[str, float],
    breakeven: float | None,
    as_at: date,
) -> Workbook:
    """schedule: {month_key: {week_no: [Group,...]}} from dispatch_plan_schedule.schedule_groups().
    invoiced_to_date_by_month: only the real current calendar month should
    have a non-zero value here -- future months haven't invoiced anything
    yet."""
    wb = Workbook()
    overview_ws = wb.active
    overview_ws.title = 'Overview'

    months_in_order = sorted(schedule.keys(), key=month_from_key)
    used_titles: set[str] = set()
    month_refs = []

    for month_key in months_in_order:
        title = _sheet_title(month_key, used_titles)
        ws = wb.create_sheet(title)
        refs = _write_month_sheet(
            ws, month_key, schedule[month_key], monthly_target,
            invoiced_to_date_by_month.get(month_key, 0.0), breakeven, as_at,
        )
        month_refs.append((month_key, refs))

    # Holding section lives at the bottom of the LAST month sheet -- matches
    # the reference workbook (one workbook-wide "other orders" list, not
    # one per month).
    if month_refs:
        last_ws = wb[month_refs[-1][1]['sheet_title']]
        _write_holding_section(last_ws, last_ws.max_row + 2, holding)

    _write_overview_sheet(overview_ws, month_refs)
    return wb


def _write_overview_sheet(ws: Worksheet, month_refs: list[tuple[str, dict]]):
    ws.column_dimensions['A'].width = 22
    for col, width in zip('BCDEF', (16, 18, 18, 16, 14)):
        ws.column_dimensions[col].width = width

    headers = ['Month', 'Target', 'Invoiced to date', 'Total projected sales', 'Variance to target', 'On track?']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = BOLD

    row = 2
    for month_key, refs in month_refs:
        title = refs['sheet_title']
        ref = f"'{title}'" if ' ' in title else title
        ws.cell(row=row, column=1, value=title)
        for col, cell_ref in zip((2, 3, 4), (refs['target_cell'], refs['invoiced_cell'], refs['total_projected_cell'])):
            c = ws.cell(row=row, column=col, value=f'={ref}!{cell_ref}')
            c.number_format = CURRENCY_FMT
        variance_col = 5
        c = ws.cell(row=row, column=variance_col, value=f'=D{row}-B{row}')
        c.number_format = CURRENCY_FMT
        c = ws.cell(row=row, column=6, value=f'=IF(E{row}>=0,"ON TRACK","BEHIND")')
        row += 1
